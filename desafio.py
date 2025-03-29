import csv #Permite a exportação dos dados para o formato especificado pelo desafio.
import re #Permite a utilização de expressões específicas.
from openpyxl import load_workbook #Permite a importação de arquivos xlsx.
from datetime import datetime #Permite manipulação de dados em formato de data e hora.

#Pequeno guia para substituição das siglas dos Estados brasileiros por seus nomes completos.
siglas_estados = {
    "AC": "Acre", "AL": "Alagoas", "AP": "Amapa", "AM": "Amazonas", "BA": "Bahia", "CE": "Ceara", "DF": "Distrito Federal", "ES": "Espirito Santo", "GO": "Goias", "MA": "Maranhao", "MT": "Mato Grosso", "MS": "Mato Grosso do Sul", "MG": "Minas Gerais", "PA": "Para", "PB": "Paraiba", "PR": "Parana", "PE": "Pernambuco", "PI": "Piaui", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte", "RS": "Rio Grande do Sul", "RO": "Rondonia", "RR": "Roraima", "SC": "Santa Catarina", "SP": "Sao Paulo", "SE": "Sergipe", "TO": "Tocantins"
}

#Pequeno guia para a substituição dos meses em números para os meses escritos por extenso.
meses = {
    "01": "Janeiro", "02": "Fevereiro", "03": "Marco", "04": "Abril", "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto", "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
}

#Dizendo pro Python de qual arquivo iremos importar as informações para tratamento (Utilizando a biblioteca openpyxl importada anteriormente).
wb = load_workbook(filename='arquivoEntrada.xlsx')
sheet = wb.active

#Iniciando a lista que vai armazenar os dados a serem convertidos para CSV.
csv_data = []
first_row = True

#Função para fazer com que os valores de dinheiro tenham duas casas depois da vírgula.
def formatar_numero_com_virgula(numero):
    try:
        numero_formatado = float(numero) / 100
        return "{:,.2f}".format(numero_formatado).replace(",", ";").replace(".", ",")
    except ValueError:
        return numero

for row in sheet.iter_rows(values_only=True):
    row = list(row)

    #Função que verifica se as células da segunda coluna contêm texto.
    if len(row) > 1 and isinstance(row[1], str):
        if first_row:

            #Correção do título da coluna 'Numero Incricao Empresa' para 'Numero Inscricao Empresa'.
            row[1] = row[1].replace("Numero Incricao Empresa", "Numero Inscricao Empresa")
            first_row = False
        else:
             
             #Substitui as siglas dos Estados pelo nome completo do Estado na coluna 7.
             for sigla, estado in siglas_estados.items():

                #Expressão re para garantir que apenas siglas isoladas sejam substituídas.
                row[6] = re.sub(r'\b' + sigla + r'\b', estado, row[6])

            #Remoção de todos os caracteres não numéricos da segunda coluna.
             row[1] = re.sub(r'\D', '', row[1])

        #Formatação das datas, de '2024-10-03 0:00:00' para '10/Marco/2024'.
        for i in range(len(row)):
         if isinstance(row[i], datetime):
            row[i] = row[i].strftime("%m/%d/%Y")
            dia, mes, ano = row[i].split("/")
            row[i] = f"{dia}/{meses[mes]}/{ano}"

        #Alterando a formatação padrão do Python para colocar uma vírgula ao invés de um ponto como separador na coluna 11.
        if len(row) > 10 and isinstance(row[10], (int, float, str)):
         row[10] = formatar_numero_com_virgula(row[10])

    #Adição das linhas modificadas na lista de dados.
    csv_data.append(row)

#Ordenação os dados em ordem numérica de acordo com os números inseridos na coluna 8.
csv_data[1:] = sorted(csv_data[1:], key=lambda x: (x[7] if isinstance(x[7], (int, float)) else float('inf')))

#Gravação dos dados formatados no arquivo final.
with open('SampleData.csv', 'w', newline='') as csv_saida:
    writer = csv.writer(csv_saida, delimiter=';')
    writer.writerows(csv_data)